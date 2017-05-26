'''
######
#	Module for creating and saving an Excel file
######
''' 

from openpyxl import Workbook
from openpyxl.formatting import Rule
from openpyxl.styles import Font, PatternFill, Border, Alignment
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
from openpyxl.cell import get_column_letter

from log import Log

class CreateExcel:
	######
	#	Logging is optional. If bool is true, then log
	######
	def __init__(self, log_path=None, log_name=None):
		######
		#	Create Workbook object, as well as worksheet
		######
		self.wb = Workbook()
		self.ws = self.wb.active
		self.text_attribute = Alignment(horizontal="center", vertical="bottom",
								text_rotation=0, wrap_text=False,
								shrink_to_fit=False, indent=0
							)
		self.bg_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
		
		######
		#	Track current Excel row
		######
		self.round_number = 1
		
		if log_name is not None:
			self.log_name = log_name
			self.log = Log(log_path, log_name)
		else:
			self.log_name = None
		
		self.caret_separater = "^" * 11

	######
	#	Write to log file if a log name was
	#	provided when object was initialized
	#	1 = DEBUG
	#	2 = INFO
	#	3 = WARNING
	#	4 = ERROR
	#	5 = CRITICAL
	######
	def update_log(self, m, level=None):
		if self.log_name is not None:
			if level is None:
				self.log.log_error(m + 
									"\n" + 
									self.caret_separater + 
									" Log level int not specified. Defaulted to ERROR."
									)
			elif level == 1:
				self.log.log_debug(m)
			elif level == 2:
				self.log.log_info(m)
			elif level == 3:
				self.log.log_warning(m)
			elif level == 4:
				self.log.log_error(m)
			elif level == 5:
				self.log.log_critical(m)
			else:
				self.log.log_error(m +
									"\n" +
									self.caret_separater +
									" Log level int %s isn't supported. Defaulted to ERROR."
									% str(level)
									)
			
	######
	#	Change header dimensions
	######
	def update_cell_dimensions(self, height, *width):
		for x in range(len(width)):
			self.update_log("Obtaining column letter and updating its width.", 2)
			
			column = get_column_letter(x+1)
			self.ws.column_dimensions[column].width = width[x]
			
			self.update_log("Updating column " + column + " to width of " + str(width[x]), 2)
				
		self.update_log("Updating row height.", 2)
		self.ws.row_dimensions[self.round_number].height = height
		self.update_log("Row height is now " + str(height), 2)
	
	######
	#	Input cell values and format
	######
	def update_cell(self, is_header=None, *contents):
		if is_header is None:
			self.update_log("is_header is None, and it needs to be specified as True or False.", 5)
			
			print("Please specify True or False as to whether this update is a header.")
			return False
		cell_content = contents
		cells = []
		
		######
		#	Create and append a cell object to cells list
		#	Number of cell objects appended contingent on
		#	number of arguments provided
		#	Afterwards, increment round_number to move to next row
		######
		for x in range(len(cell_content)):
			self.update_log("Updating cell content. is_header = %s" % is_header, 2)
			
			cells.append(self.ws.cell(row=self.round_number, column=x+1, value=cell_content[x]))
			self.update_log("Appending %s to cells[] and content is %s." % (cells[x], cell_content[x]), 2)
			
			cells[x].alignment = self.text_attribute
			self.update_log("Updating %s alignment to %s." % (cells[x], self.text_attribute), 2)
			
			if is_header:
				self.update_log("is_header is %s." % is_header, 2)
				cells[x].fill = self.bg_fill
				self.update_log("Updating header %s fill to %s." % (cells[x], self.bg_fill), 2)
		self.round_number += 1
		self.update_log("round_number has been incremented by 1.", 2)

	def save_file(self, file_path=None, file_name=None): 
		if file_path is None:
			self.update_log("File path was not specified.", 5)
			
			print("\nPlease specify the file path.")
			return False
		
		if file_name is None:
			self.update_log("File name was not specified.", 5)
			
			print("\nPlease specify the file name.")
			return False
	
		######
		#	Save file
		######
		try:
			self.wb.save(file_path + "/" + file_name)
		except IOError:
			self.update_log("IOError when saving file. Retrying without '/' after file path.", 3)
			
			print("Trying without '/' after file path.")
			try:
				self.wb.save(file_path + file_name)
			except IOError:
				self.update_log("IOError when saving file. File saved unsuccessfully", 5)
				
				print("\nX----X File saved unsuccessfully.")
				return False
		self.update_log("Log file saved successfully", 2)
		print("\n!----! Log file saved successfully.")
